import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = ROOT / "Data_Tanaman_Padi_Sumatera_siap_metabase.xlsx"
OUTPUT_PATH = ROOT / "public" / "data" / "padi-sumatera.json"
SHEET_NAME = "Data_Siap_Metabase"

FIELD_MAP = {
    "Provinsi": "province",
    "Tahun": "year",
    "Produksi": "production",
    "Luas Panen": "harvestArea",
    "Curah hujan": "rainfall",
    "Kelembapan": "humidity",
    "Suhu rata-rata": "avgTemperature",
    "Negara": "country",
    "Ibu Kota Provinsi": "capitalCity",
    "Latitude": "latitude",
    "Longitude": "longitude",
    "Lokasi Pin": "pinLocation",
    "Jenis Koordinat": "coordinateType",
    "Sumber Koordinat": "coordinateSource",
}

NUMERIC_FIELDS = {
    "year",
    "production",
    "harvestArea",
    "rainfall",
    "humidity",
    "avgTemperature",
    "latitude",
    "longitude",
}


def slugify(value):
    value = unicodedata.normalize("NFD", str(value).lower())
    value = "".join(char for char in value if unicodedata.category(char) != "Mn")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")


def to_number(value):
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return value
    return float(str(value).replace(",", "."))


def main():
    workbook = load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    sheet = workbook[SHEET_NAME]
    headers = [cell.value for cell in sheet[4]]
    records = []

    for row in sheet.iter_rows(min_row=5, values_only=True):
      if not any(row):
          continue

      record = {}
      for header, value in zip(headers, row):
          field = FIELD_MAP.get(header)
          if not field:
              continue

          record[field] = to_number(value) if field in NUMERIC_FIELDS else value

      record["id"] = f"{slugify(record['province'])}-{int(record['year'])}"
      records.append(record)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps(records, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Wrote {len(records)} records to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
